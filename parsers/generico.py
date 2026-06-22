"""Parser genérico — fallback cuando no se detecta el banco."""
import re
import pdfplumber
from openpyxl import load_workbook
from parsers.base import BaseParser


class GenericoPdfParser(BaseParser):

    def parsear(self, ruta: str, password: str = None) -> dict:
        res = self.respuesta_vacia()
        movimientos = []
        errores = ["Banco no detectado automáticamente. Usando parser genérico."]
        try:
            open_kwargs = {"password": password} if password else {}
            with pdfplumber.open(ruta, **open_kwargs) as pdf:
                for page in pdf.pages:
                    tablas = page.extract_tables()
                    if tablas:
                        for tabla in tablas:
                            for fila in tabla:
                                mov = self._fila_a_mov(fila)
                                if mov:
                                    movimientos.append(mov)
                    else:
                        texto = page.extract_text(layout=True) or ""
                        movimientos.extend(self._parsear_texto(texto))
        except Exception as e:
            errores.append(f"Error PDF genérico: {str(e)}")
        if not res["cabecera"].get("periodo") and movimientos:
            res["cabecera"]["periodo"] = self.periodo_desde_fechas(movimientos)
        res.update({"movimientos": movimientos, "errores": errores, "total_leidos": len(movimientos)})
        return res

    def _fila_a_mov(self, fila: list) -> dict | None:
        if not fila or len(fila) < 3:
            return None
        cols = [str(c or "").strip() for c in fila]
        # Buscar columna con fecha
        fecha = None
        fecha_col = -1
        for i, col in enumerate(cols):
            f = self.parse_fecha(col)
            if f:
                fecha = f
                fecha_col = i
                break
        if not fecha:
            return None
        # Buscar números como importes
        importes = [(i, self.parse_importe(c)) for i, c in enumerate(cols) if i > fecha_col and self.parse_importe(c) > 0]
        if len(importes) < 1:
            return None
        desc = " ".join(c for i, c in enumerate(cols) if i > fecha_col and self.parse_importe(c) == 0 and c)
        if not desc:
            return None
        saldo   = importes[-1][1] if len(importes) >= 2 else None
        importe = importes[-2][1] if len(importes) >= 2 else importes[0][1]
        return {
            "fecha_operacion": fecha, "fecha_valor": fecha, "referencia": None,
            "descripcion": desc[:300], "tipo": "cargo", "importe": importe,
            "saldo_banco": saldo, "moneda": "PEN", "tipo_cambio": 1.0,
        }

    def _parsear_texto(self, texto: str) -> list:
        movs = []
        for linea in texto.splitlines():
            m = re.match(r'^(\d{2}/\d{2}/\d{4})\s+', linea)
            if not m:
                continue
            fecha = self.parse_fecha(m.group(1))
            if not fecha:
                continue
            importes = re.findall(r'[\d]{1,3}(?:[,\.]\d{3})*[,\.]\d{2}', linea)
            if len(importes) < 2:
                continue
            saldo   = self.parse_importe(importes[-1])
            importe = self.parse_importe(importes[-2])
            desc    = re.sub(r'\d{2}/\d{2}/\d{4}', '', linea)
            desc    = re.sub(r'[\d,\.]+', '', desc).strip()
            if not desc or importe <= 0:
                continue
            movs.append({
                "fecha_operacion": fecha, "fecha_valor": fecha, "referencia": None,
                "descripcion": desc[:300], "tipo": "cargo", "importe": importe,
                "saldo_banco": saldo, "moneda": "PEN", "tipo_cambio": 1.0,
            })
        return movs


class GenericoExcelParser(BaseParser):

    def parsear(self, ruta: str) -> dict:
        res = self.respuesta_vacia()
        movimientos = []
        errores = ["Banco no detectado. Usando parser genérico Excel."]
        try:
            wb = load_workbook(ruta, read_only=True, data_only=True)
            ws = wb.active
            # Detectar fila inicio y columnas automáticamente
            fila_inicio, col_map = self._detectar_estructura(ws)
            for row in ws.iter_rows(min_row=fila_inicio, values_only=True):
                if not row:
                    continue
                fecha = self.parse_fecha(row[col_map["fecha"]])
                if not fecha:
                    continue
                desc  = str(row[col_map["desc"]] or "").strip() if col_map.get("desc") is not None else ""
                cargo = self.parse_importe(row[col_map["cargo"]] if col_map.get("cargo") is not None else None)
                abono = self.parse_importe(row[col_map["abono"]] if col_map.get("abono") is not None else None)
                saldo = self.parse_importe(row[col_map["saldo"]] if col_map.get("saldo") is not None else None)
                if cargo <= 0 and abono <= 0 or not desc:
                    continue
                tipo = "cargo" if cargo > 0 else "abono"
                movimientos.append({
                    "fecha_operacion": fecha, "fecha_valor": fecha, "referencia": None,
                    "descripcion": desc[:300], "tipo": tipo,
                    "importe": cargo if cargo > 0 else abono,
                    "saldo_banco": saldo or None, "moneda": "PEN", "tipo_cambio": 1.0,
                })
            wb.close()
        except Exception as e:
            errores.append(f"Error Excel genérico: {str(e)}")
        if not res["cabecera"].get("periodo") and movimientos:
            res["cabecera"]["periodo"] = self.periodo_desde_fechas(movimientos)
        res.update({"movimientos": movimientos, "errores": errores, "total_leidos": len(movimientos)})
        return res

    def _detectar_estructura(self, ws):
        """Detecta fila de inicio y mapa de columnas buscando encabezados."""
        col_map = {"fecha": 0, "desc": 2, "cargo": 3, "abono": 4, "saldo": 5}
        fila_inicio = 5
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            if not row:
                continue
            textos = [str(c or "").upper() for c in row]
            if any("FECHA" in t for t in textos):
                for j, t in enumerate(textos):
                    if "FECHA" in t:
                        col_map["fecha"] = j
                    elif "DESCRIPCI" in t or "CONCEPTO" in t or "GLOSA" in t:
                        col_map["desc"] = j
                    elif "CARGO" in t or "DEBITO" in t or "DÉBITO" in t:
                        col_map["cargo"] = j
                    elif "ABONO" in t or "CREDITO" in t or "CRÉDITO" in t:
                        col_map["abono"] = j
                    elif "SALDO" in t:
                        col_map["saldo"] = j
                fila_inicio = i + 1
                break
        return fila_inicio, col_map
