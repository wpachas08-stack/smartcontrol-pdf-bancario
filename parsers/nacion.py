"""Banco de la Nación Perú — PDF y Excel"""
import re
import pdfplumber
from openpyxl import load_workbook
from parsers.base import BaseParser


class NacionPdfParser(BaseParser):
    """
    Banco de la Nación emite PDFs con formato tabular simple.
    Columnas: Fecha | Operación | Descripción | Débito | Crédito | Saldo
    """

    def parsear(self, ruta: str) -> dict:
        res = self.respuesta_vacia()
        movimientos = []
        errores = []
        try:
            with pdfplumber.open(ruta) as pdf:
                texto_cab = ""
                for i, page in enumerate(pdf.pages):
                    tablas = page.extract_tables()
                    if tablas:
                        for tabla in tablas:
                            for fila in tabla:
                                mov = self._fila_a_mov(fila)
                                if mov:
                                    movimientos.append(mov)
                    else:
                        texto = page.extract_text(layout=True) or ""
                        if i == 0:
                            texto_cab = texto
                        movimientos.extend(self._parsear_texto(texto))
                res["cabecera"] = self._extraer_cabecera(texto_cab)
        except Exception as e:
            errores.append(f"Error PDF Banco de la Nación: {str(e)}")
        if not res["cabecera"].get("periodo") and movimientos:
            res["cabecera"]["periodo"] = self.periodo_desde_fechas(movimientos)
        res.update({"movimientos": movimientos, "errores": errores, "total_leidos": len(movimientos)})
        return res

    def _fila_a_mov(self, fila: list) -> dict | None:
        if not fila or len(fila) < 4:
            return None
        cols = [str(c or "").strip() for c in fila]
        fecha = self.parse_fecha(cols[0])
        if not fecha:
            return None
        # BN: Fecha | Nro Op | Descripcion | Debito | Credito | Saldo
        ref    = cols[1] if len(cols) > 1 else ""
        desc   = cols[2] if len(cols) > 2 else ""
        debito = self.parse_importe(cols[3] if len(cols) > 3 else "")
        credito= self.parse_importe(cols[4] if len(cols) > 4 else "")
        saldo  = self.parse_importe(cols[5] if len(cols) > 5 else "")
        if debito <= 0 and credito <= 0:
            return None
        tipo = "cargo" if debito > 0 else "abono"
        return {
            "fecha_operacion": fecha, "fecha_valor": fecha,
            "referencia": ref or None, "descripcion": desc[:300],
            "tipo": tipo, "importe": debito if debito > 0 else credito,
            "saldo_banco": saldo or None, "moneda": "PEN", "tipo_cambio": 1.0,
        }

    def _parsear_texto(self, texto: str) -> list:
        movs = []
        for linea in texto.splitlines():
            m = re.match(r'^(\d{2}/\d{2}/\d{4})\s+', linea)
            if not m:
                continue
            fecha = self.parse_fecha(m.group(1))
            if not fecha:
                continue
            importes = re.findall(r'[\d]{1,3}(?:[,\.]\d{3})*[,\.]\d{2}', linea)
            if len(importes) < 2:
                continue
            saldo   = self.parse_importe(importes[-1])
            importe = self.parse_importe(importes[-2])
            desc    = re.sub(r'\d{2}/\d{2}/\d{4}', '', linea)
            desc    = re.sub(r'[\d,\.]+', '', desc).strip()
            if not desc or importe <= 0:
                continue
            movs.append({
                "fecha_operacion": fecha, "fecha_valor": fecha, "referencia": None,
                "descripcion": desc[:300], "tipo": "cargo", "importe": importe,
                "saldo_banco": saldo, "moneda": "PEN", "tipo_cambio": 1.0,
            })
        return movs

    def _extraer_cabecera(self, texto: str) -> dict:
        t = texto.upper()
        cab = {"numero_cuenta": None, "periodo": None, "saldo_inicial": None, "saldo_final": None, "moneda": "PEN"}
        m = re.search(r'(?:CUENTA|CTA)[^\d]+([\d\s\-]{6,25})', t)
        if m:
            cab["numero_cuenta"] = re.sub(r'[\s\-]', '', m.group(1))[:20]
        m = re.search(r'(\d{2})/(\d{4})', t)
        if m:
            cab["periodo"] = m.group(2) + m.group(1)
        return cab


class NacionExcelParser(BaseParser):

    def parsear(self, ruta: str) -> dict:
        res = self.respuesta_vacia()
        movimientos = []
        errores = []
        try:
            wb = load_workbook(ruta, read_only=True, data_only=True)
            ws = wb.active
            fila_inicio = self._detectar_inicio(ws)
            for row in ws.iter_rows(min_row=fila_inicio, values_only=True):
                if not row:
                    continue
                fecha = self.parse_fecha(row[0])
                if not fecha:
                    continue
                ref    = str(row[1] or "").strip() if len(row) > 1 else ""
                desc   = str(row[2] or "").strip() if len(row) > 2 else ""
                debito = self.parse_importe(row[3] if len(row) > 3 else None)
                credito= self.parse_importe(row[4] if len(row) > 4 else None)
                saldo  = self.parse_importe(row[5] if len(row) > 5 else None)
                if debito <= 0 and credito <= 0 or not desc:
                    continue
                tipo = "cargo" if debito > 0 else "abono"
                movimientos.append({
                    "fecha_operacion": fecha, "fecha_valor": fecha,
                    "referencia": ref or None, "descripcion": desc[:300],
                    "tipo": tipo, "importe": debito if debito > 0 else credito,
                    "saldo_banco": saldo or None, "moneda": "PEN", "tipo_cambio": 1.0,
                })
            wb.close()
        except Exception as e:
            errores.append(f"Error Excel BN: {str(e)}")
        if not res["cabecera"].get("periodo") and movimientos:
            res["cabecera"]["periodo"] = self.periodo_desde_fechas(movimientos)
        res.update({"movimientos": movimientos, "errores": errores, "total_leidos": len(movimientos)})
        return res

    def _detectar_inicio(self, ws) -> int:
        for i, row in enumerate(ws.iter_rows(min_row=3, max_row=20, values_only=True), start=3):
            if row and self.parse_fecha(row[0]):
                return i
        return 5
