"""Interbank Perú — PDF y Excel"""
import re
import pdfplumber
from openpyxl import load_workbook
from parsers.base import BaseParser


class InterbankPdfParser(BaseParser):

    def parsear(self, ruta: str, password: str = None) -> dict:
        res = self.respuesta_vacia()
        movimientos = []
        errores = []
        try:
            open_kwargs = {"password": password} if password else {}
            with pdfplumber.open(ruta, **open_kwargs) as pdf:
                for page in pdf.pages:
                    tablas = page.extract_tables()
                    if tablas:
                        for tabla in tablas:
                            for fila in tabla:
                                mov = self._fila_a_mov(fila)
                                if mov:
                                    movimientos.append(mov)
                    else:
                        texto = page.extract_text(layout=True) or ""
                        movimientos.extend(self._parsear_texto(texto))
        except Exception as e:
            errores.append(f"Error PDF Interbank: {str(e)}")
        if not res["cabecera"].get("periodo") and movimientos:
            res["cabecera"]["periodo"] = self.periodo_desde_fechas(movimientos)
        res.update({"movimientos": movimientos, "errores": errores, "total_leidos": len(movimientos)})
        return res

    def _fila_a_mov(self, fila: list) -> dict | None:
        if not fila or len(fila) < 4:
            return None
        cols = [str(c or "").strip() for c in fila]
        fecha = self.parse_fecha(cols[0])
        if not fecha:
            return None
        # Interbank: Fecha | FechaValor | Canal | Descripcion | Ref | Cargos | Abonos | Saldo
        fecha_val = self.parse_fecha(cols[1]) if len(cols) > 1 else fecha
        canal     = cols[2] if len(cols) > 2 else ""
        desc      = cols[3] if len(cols) > 3 else ""
        ref       = cols[4] if len(cols) > 4 else ""
        cargo     = self.parse_importe(cols[5] if len(cols) > 5 else "")
        abono     = self.parse_importe(cols[6] if len(cols) > 6 else "")
        saldo     = self.parse_importe(cols[7] if len(cols) > 7 else "")
        if cargo <= 0 and abono <= 0:
            return None
        desc_full = f"{canal} {desc}".strip() if canal else desc
        tipo = "cargo" if cargo > 0 else "abono"
        return {
            "fecha_operacion": fecha, "fecha_valor": fecha_val or fecha,
            "referencia": ref or None, "descripcion": desc_full[:300],
            "tipo": tipo, "importe": cargo if cargo > 0 else abono,
            "saldo_banco": saldo or None, "moneda": "PEN", "tipo_cambio": 1.0,
        }

    def _parsear_texto(self, texto: str) -> list:
        movs = []
        for linea in texto.splitlines():
            m = re.match(r'^(\d{2}/\d{2}/\d{4})\s+', linea)
            if not m:
                continue
            fecha = self.parse_fecha(m.group(1))
            if not fecha:
                continue
            importes = re.findall(r'[\d]{1,3}(?:[,\.]\d{3})*[,\.]\d{2}', linea)
            if len(importes) < 2:
                continue
            saldo   = self.parse_importe(importes[-1])
            importe = self.parse_importe(importes[-2])
            desc = re.sub(r'\d{2}/\d{2}/\d{4}', '', linea)
            desc = re.sub(r'[\d,\.]+', '', desc).strip()
            if not desc or importe <= 0:
                continue
            movs.append({
                "fecha_operacion": fecha, "fecha_valor": fecha, "referencia": None,
                "descripcion": desc[:300], "tipo": "cargo", "importe": importe,
                "saldo_banco": saldo, "moneda": "PEN", "tipo_cambio": 1.0,
            })
        return movs


class InterbankExcelParser(BaseParser):

    def parsear(self, ruta: str) -> dict:
        res = self.respuesta_vacia()
        movimientos = []
        errores = []
        try:
            wb = load_workbook(ruta, read_only=True, data_only=True)
            ws = wb.active
            fila_inicio = self._detectar_inicio(ws)
            for row in ws.iter_rows(min_row=fila_inicio, values_only=True):
                if not row:
                    continue
                fecha = self.parse_fecha(row[0])
                if not fecha:
                    continue
                fecha_val = self.parse_fecha(row[1]) if len(row) > 1 else fecha
                canal     = str(row[2] or "").strip() if len(row) > 2 else ""
                desc      = str(row[3] or "").strip() if len(row) > 3 else ""
                ref       = str(row[4] or "").strip() if len(row) > 4 else ""
                cargo     = self.parse_importe(row[5] if len(row) > 5 else None)
                abono     = self.parse_importe(row[6] if len(row) > 6 else None)
                saldo     = self.parse_importe(row[7] if len(row) > 7 else None)
                if cargo <= 0 and abono <= 0 or not desc:
                    continue
                tipo = "cargo" if cargo > 0 else "abono"
                desc_full = f"{canal} {desc}".strip() if canal else desc
                movimientos.append({
                    "fecha_operacion": fecha, "fecha_valor": fecha_val or fecha,
                    "referencia": ref or None, "descripcion": desc_full[:300],
                    "tipo": tipo, "importe": cargo if cargo > 0 else abono,
                    "saldo_banco": saldo or None, "moneda": "PEN", "tipo_cambio": 1.0,
                })
            wb.close()
        except Exception as e:
            errores.append(f"Error Excel Interbank: {str(e)}")
        if not res["cabecera"].get("periodo") and movimientos:
            res["cabecera"]["periodo"] = self.periodo_desde_fechas(movimientos)
        res.update({"movimientos": movimientos, "errores": errores, "total_leidos": len(movimientos)})
        return res

    def _detectar_inicio(self, ws) -> int:
        for i, row in enumerate(ws.iter_rows(min_row=3, max_row=20, values_only=True), start=3):
            if row and self.parse_fecha(row[0]):
                return i
        return 6
