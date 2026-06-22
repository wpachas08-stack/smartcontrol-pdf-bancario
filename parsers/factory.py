"""
ParserFactory — detecta el banco y retorna el parser correcto.
Desencripta el PDF antes de pasarlo al parser.
"""
import re
import tempfile
import os
from pathlib import Path

from pypdf import PdfReader, PdfWriter
import pdfplumber


# Palabras clave para detectar banco desde el texto del PDF/Excel
FIRMAS_BANCO = {
    "bcp":       ["BANCO DE CREDITO", "BANCO DE CRÉDITO", "CREDIBANCO", " BCP "],
    "bbva":      ["BBVA", "BANCO CONTINENTAL", "CONTINENTAL"],
    "scotiabank":["SCOTIABANK", "BANCO SCOTIABANK"],
    "interbank": ["INTERBANK", "BANCO INTERNACIONAL"],
    "nacion":    ["BANCO DE LA NACION", "BANCO DE LA NACIÓN", "BN "],
}


class ParserFactory:

    def procesar(
        self,
        ruta: str,
        extension: str,
        ruc: str,
        banco_forzado: str = "",
    ) -> dict:
        """
        Punto de entrada. Desencripta si es PDF, detecta banco, parsea.
        """
        ruta_trabajo  = ruta
        ruta_dec_tmp  = None

        try:
            if extension == "pdf":
                # Primero intentar abrir directamente (PDF ya desbloqueado)
                try:
                    with pdfplumber.open(ruta) as pdf:
                        _ = len(pdf.pages)
                    ruta_trabajo  = ruta
                    ruta_dec_tmp  = None
                except Exception:
                    # Si falla, intentar desencriptar
                    ruta_trabajo, ruta_dec_tmp = self._desencriptar(ruta, ruc)

            banco = banco_forzado.lower() if banco_forzado else \
                    self._detectar_banco(ruta_trabajo, extension)

            parser = self._get_parser(banco, extension)
            # Pasar ruc como password para que pdfplumber intente con contraseña
            result = parser.parsear(ruta_trabajo, password=ruc if extension == 'pdf' else None)
            result["ok"]    = True
            result["banco"] = banco
            return result

        finally:
            if ruta_dec_tmp and os.path.exists(ruta_dec_tmp):
                os.unlink(ruta_dec_tmp)

    # ------------------------------------------------------------------
    # Desencriptar PDF con pikepdf (soporta AES-128/256 y RC4)
    # ------------------------------------------------------------------
    def _desencriptar(self, ruta: str, password: str):
        """
        Desencripta PDF protegido. BCP usa formato propietario con RUC como contraseña.
        Usa qpdf del sistema operativo (disponible en Render/Linux).
        """
        fd, ruta_dec = tempfile.mkstemp(suffix=".pdf", prefix="dec_")
        os.close(fd)

        # Método 1: qpdf (disponible en Render/Linux — maneja formato BCP)
        try:
            import subprocess
            result = subprocess.run(
                ['qpdf', '--decrypt', f'--password={password}', ruta, ruta_dec],
                capture_output=True, text=True, timeout=30
            )
            if result.returncode == 0 and os.path.exists(ruta_dec) and os.path.getsize(ruta_dec) > 0:
                # Verificar que es legible
                try:
                    with pdfplumber.open(ruta_dec) as pdf:
                        _ = len(pdf.pages)
                    return ruta_dec, ruta_dec
                except Exception:
                    pass
        except Exception:
            pass

        # Método 2: pdfplumber con password directamente (sin desencriptar a disco)
        try:
            with pdfplumber.open(ruta, password=password) as pdf:
                _ = len(pdf.pages)
            # Funciona con password directa — retornar original, el parser usará password
            if os.path.exists(ruta_dec):
                os.unlink(ruta_dec)
            return ruta, None
        except Exception:
            pass

        # Método 3: pypdf
        try:
            reader = PdfReader(ruta)
            if reader.is_encrypted:
                reader.decrypt(password)
            writer = PdfWriter()
            for page in reader.pages:
                writer.add_page(page)
            with open(ruta_dec, 'wb') as f:
                writer.write(f)
            return ruta_dec, ruta_dec
        except Exception:
            pass

        if os.path.exists(ruta_dec):
            os.unlink(ruta_dec)
        raise ValueError(
            f"No se pudo desencriptar el PDF con RUC '{password}'. "
            "Verifique que el RUC coincide con el estado de cuenta."
        )

    # ------------------------------------------------------------------
    # Detección de banco
    # ------------------------------------------------------------------
    def _detectar_banco(self, ruta: str, extension: str) -> str:
        texto = ""
        try:
            if extension == "pdf":
                with pdfplumber.open(ruta) as pdf:
                    for page in pdf.pages[:2]:
                        texto += (page.extract_text() or "")
                        if len(texto) > 500:
                            break
            else:
                from openpyxl import load_workbook
                wb = load_workbook(ruta, read_only=True, data_only=True)
                ws = wb.active
                for row in ws.iter_rows(max_row=15, values_only=True):
                    for cell in row:
                        if cell:
                            texto += str(cell) + " "
                wb.close()
        except Exception:
            pass

        texto_up = texto.upper()
        for banco, firmas in FIRMAS_BANCO.items():
            for firma in firmas:
                if firma in texto_up:
                    return banco
        return "generico"

    # ------------------------------------------------------------------
    # Instanciar parser
    # ------------------------------------------------------------------
    def _get_parser(self, banco: str, extension: str):
        from parsers.bcp        import BcpPdfParser, BcpExcelParser
        from parsers.bbva       import BbvaPdfParser, BbvaExcelParser
        from parsers.scotiabank import ScotiabankPdfParser, ScotiabankExcelParser
        from parsers.interbank  import InterbankPdfParser, InterbankExcelParser
        from parsers.nacion     import NacionPdfParser, NacionExcelParser
        from parsers.generico   import GenericoPdfParser, GenericoExcelParser

        mapa = {
            ("bcp",        "pdf"):  BcpPdfParser,
            ("bcp",        "xlsx"): BcpExcelParser,
            ("bcp",        "xls"):  BcpExcelParser,
            ("bbva",       "pdf"):  BbvaPdfParser,
            ("bbva",       "xlsx"): BbvaExcelParser,
            ("bbva",       "xls"):  BbvaExcelParser,
            ("scotiabank", "pdf"):  ScotiabankPdfParser,
            ("scotiabank", "xlsx"): ScotiabankExcelParser,
            ("scotiabank", "xls"):  ScotiabankExcelParser,
            ("interbank",  "pdf"):  InterbankPdfParser,
            ("interbank",  "xlsx"): InterbankExcelParser,
            ("interbank",  "xls"):  InterbankExcelParser,
            ("nacion",     "pdf"):  NacionPdfParser,
            ("nacion",     "xlsx"): NacionExcelParser,
            ("nacion",     "xls"):  NacionExcelParser,
        }
        key   = (banco, extension)
        clase = mapa.get(key)
        if clase is None:
            clase = GenericoPdfParser if extension == "pdf" else GenericoExcelParser
        return clase()
