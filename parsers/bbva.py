"""
Parser BBVA Perú (Continental)

PDF: Layout con encabezados variables, columnas: Fecha | Fecha Valor | Referencia | Descripción | Cargos | Abonos | Saldo
Excel: Fila ~10 inicio, col A=Fecha, B=Fecha Valor, C=Referencia, D=Descripción, E=Cargos, F=Abonos, G=Saldo
"""
import re
import pdfplumber
from openpyxl import load_workbook
from parsers.base import BaseParser


class BbvaPdfParser(BaseParser):

    def parsear(self, ruta: str) -> dict:
        res         = self.respuesta_vacia()
        movimientos = []
        errores     = []

        try:
            with pdfplumber.open(ruta) as pdf:
                texto_completo = ""
                for page in pdf.pages:
                    # BBVA responde bien a extracción por tabla
                    tablas = page.extract_tables()
                    if tablas:
                        for tabla in tablas:
                            for fila in tabla:
                                mov = self._fila_a_mov(fila)
                                if mov:
                                    movimientos.append(mov)
                    else:
                        texto = page.extract_text(layout=True) or ""
                        texto_completo += texto
                        movimientos.extend(self._parsear_texto(texto))

                if texto_completo:
                    res["cabecera"] = self._extraer_cabecera(texto_completo)

        except Exception as e:
            errores.append(f"Error PDF BBVA: {str(e)}")

        if not res["cabecera"].get("periodo") and movimientos:
            res["cabecera"]["periodo"] = self.periodo_desde_fechas(movimientos)

        res.update({"movimientos": movimientos, "errores": errores, "total_leidos": len(movimientos)})
        return res

    def _fila_a_mov(self, fila: list) -> dict | None:
        if not fila or len(fila) < 4:
            return None
        cols = [str(c or "").strip() for c in fila]
        fecha = self.parse_fecha(cols[0])
        if not fecha:
            return None
        # BBVA: col 0=Fecha, 1=FechaValor, 2=Ref, 3=Desc, 4=Cargo, 5=Abono, 6=Saldo
        desc    = cols[3] if len(cols) > 3 else cols[2]
        cargo   = self.parse_importe(cols[4] if len(cols) > 4 else "")
        abono   = self.parse_importe(cols[5] if len(cols) > 5 else "")
        saldo   = self.parse_importe(cols[6] if len(cols) > 6 else "")
        ref     = cols[2] if len(cols) > 2 else None
        if cargo <= 0 and abono <= 0:
            return None
        tipo = "cargo" if cargo > 0 else "abono"
        return {
            "fecha_operacion": fecha, "fecha_valor": self.parse_fecha(cols[1]) or fecha,
            "referencia": ref or None, "descripcion": desc[:300],
            "tipo": tipo, "importe": cargo if cargo > 0 else abono,
            "saldo_banco": saldo or None, "moneda": "PEN", "tipo_cambio": 1.0,
        }

    def _parsear_texto(self, texto: str) -> list:
        movs = []
        for linea in texto.splitlines():
            m = re.match(r'^(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})?\s*', linea)
            if not m:
                continue
            fecha = self.parse_fecha(m.group(1))
            if not fecha:
                continue
            importes = re.findall(r'[\d]{1,3}(?:[,\.]\d{3})*[,\.]\d{2}', linea[m.end():])
            if len(importes) < 2:
                continue
            saldo   = self.parse_importe(importes[-1])
            importe = self.parse_importe(importes[-2])
            desc_raw = linea[m.end():]
            desc_raw = re.sub(r'[\d,\.]+', '', desc_raw).strip()
            if not desc_raw or importe <= 0:
                continue
            movs.append({
                "fecha_operacion": fecha, "fecha_valor": fecha, "referencia": None,
                "descripcion": desc_raw[:300], "tipo": "cargo", "importe": importe,
                "saldo_banco": saldo, "moneda": "PEN", "tipo_cambio": 1.0,
            })
        return movs

    def _extraer_cabecera(self, texto: str) -> dict:
        t = texto.upper()
        cab = {"numero_cuenta": None, "periodo": None, "saldo_inicial": None, "saldo_final": None, "moneda": "PEN"}
        m = re.search(r'(?:CUENTA|CTA)[^\d]+([\d\s\-]{6,25})', t)
        if m:
            cab["numero_cuenta"] = re.sub(r'[\s\-]', '', m.group(1))[:20]
        m = re.search(r'(\d{2})/(\d{4})', t)
        if m:
            cab["periodo"] = m.group(2) + m.group(1)
        if "DOLAR" in t or "USD" in t:
            cab["moneda"] = "USD"
        return cab


class BbvaExcelParser(BaseParser):

    def parsear(self, ruta: str) -> dict:
        res = self.respuesta_vacia()
        movimientos = []
        errores = []
        try:
            wb = load_workbook(ruta, read_only=True, data_only=True)
            ws = wb.active
            fila_inicio = self._detectar_inicio(ws)
            for row in ws.iter_rows(min_row=fila_inicio, values_only=True):
                if not row or not row[0]:
                    continue
                fecha = self.parse_fecha(row[0])
                if not fecha:
                    continue
                fecha_val = self.parse_fecha(row[1]) if len(row) > 1 else fecha
                ref       = str(row[2] or "").strip() if len(row) > 2 else ""
                desc      = str(row[3] or "").strip() if len(row) > 3 else ""
                cargo     = self.parse_importe(row[4] if len(row) > 4 else None)
                abono     = self.parse_importe(row[5] if len(row) > 5 else None)
                saldo     = self.parse_importe(row[6] if len(row) > 6 else None)
                if cargo <= 0 and abono <= 0 or not desc:
                    continue
                tipo = "cargo" if cargo > 0 else "abono"
                movimientos.append({
                    "fecha_operacion": fecha, "fecha_valor": fecha_val or fecha,
                    "referencia": ref or None, "descripcion": desc[:300],
                    "tipo": tipo, "importe": cargo if cargo > 0 else abono,
                    "saldo_banco": saldo or None, "moneda": "PEN", "tipo_cambio": 1.0,
                })
            wb.close()
        except Exception as e:
            errores.append(f"Error Excel BBVA: {str(e)}")
        if not res["cabecera"].get("periodo") and movimientos:
            res["cabecera"]["periodo"] = self.periodo_desde_fechas(movimientos)
        res.update({"movimientos": movimientos, "errores": errores, "total_leidos": len(movimientos)})
        return res

    def _detectar_inicio(self, ws) -> int:
        for i, row in enumerate(ws.iter_rows(min_row=5, max_row=25, values_only=True), start=5):
            if row and self.parse_fecha(row[0]):
                return i
        return 10
