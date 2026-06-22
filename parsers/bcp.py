"""
Parsers BCP (Banco de Crédito del Perú)

PDF: Estado de cuenta mensual, puede estar encriptado con RUC como contraseña.
     Formato tabular con columnas: Fecha | N°Op | Descripción | Cargo | Abono | Saldo

Excel: Columnas fijas desde fila ~8:
     A=Fecha | B=N°Op | C=Descripción | D=Cargo | E=Abono | F=Saldo
"""
import re
import pdfplumber
from openpyxl import load_workbook
from parsers.base import BaseParser


class BcpPdfParser(BaseParser):

    def parsear(self, ruta: str, password: str = None) -> dict:
        res = self.respuesta_vacia()
        errores = []
        movimientos = []

        try:
            open_kwargs = {}
            if password:
                open_kwargs['password'] = password
            with pdfplumber.open(ruta, **open_kwargs) as pdf:
                texto_cabecera = ""
                todas_lineas   = []

                for i, page in enumerate(pdf.pages):
                    # Intentar extracción por tabla primero (más precisa)
                    tablas = page.extract_tables({
                        "vertical_strategy":   "lines",
                        "horizontal_strategy": "lines",
                    })

                    if tablas:
                        for tabla in tablas:
                            for fila in tabla:
                                if fila and any(fila):
                                    todas_lineas.append(fila)
                    else:
                        # Fallback: texto plano con layout
                        texto = page.extract_text(layout=True) or ""
                        if i == 0:
                            texto_cabecera = texto
                        todas_lineas.extend(
                            [{"_raw": linea} for linea in texto.splitlines()]
                        )

                res["cabecera"] = self._extraer_cabecera_pdf(texto_cabecera or "")
                movimientos     = self._procesar_lineas(todas_lineas)

        except Exception as e:
            errores.append(f"Error al leer PDF BCP: {str(e)}")

        if not res["cabecera"].get("periodo") and movimientos:
            res["cabecera"]["periodo"] = self.periodo_desde_fechas(movimientos)

        res["movimientos"]  = movimientos
        res["errores"]      = errores
        res["total_leidos"] = len(movimientos)
        return res

    def _extraer_cabecera_pdf(self, texto: str) -> dict:
        cab = {
            "numero_cuenta": None,
            "periodo":       None,
            "saldo_inicial": None,
            "saldo_final":   None,
            "moneda":        "PEN",
        }
        t = texto.upper()

        # Número de cuenta
        m = re.search(r'(?:CUENTA|CTA\.?)[:\s#]+(\d[\d\s\-]{6,20})', t)
        if m:
            cab["numero_cuenta"] = re.sub(r'[\s\-]', '', m.group(1))

        # Período
        m = re.search(r'(\d{2})[/\-](\d{4})', t)
        if m:
            cab["periodo"] = m.group(2) + m.group(1).zfill(2)

        # Saldos
        m = re.search(r'SALDO\s+INICIAL[^\d]+([\d,\.]+)', t)
        if m:
            cab["saldo_inicial"] = self.parse_importe(m.group(1))

        m = re.search(r'SALDO\s+FINAL[^\d]+([\d,\.]+)', t)
        if m:
            cab["saldo_final"] = self.parse_importe(m.group(1))

        # Moneda
        if "DOLAR" in t or "USD" in t:
            cab["moneda"] = "USD"

        return cab

    def _procesar_lineas(self, lineas: list) -> list:
        """
        Procesa líneas que pueden ser listas (de tabla) o dicts (de texto plano).
        """
        movimientos = []
        en_datos    = False

        for linea in lineas:
            # Caso 1: linea es una fila de tabla (lista)
            if isinstance(linea, list):
                # Detectar fila de cabecera de tabla
                textos = [str(c or "").upper().strip() for c in linea]
                if any("FECHA" in t for t in textos) and \
                   any(t in ("CARGO", "CARGOS", "DÉBITOS", "DEBITOS") for t in textos):
                    en_datos = True
                    continue

                if not en_datos:
                    continue

                # Detectar fin de tabla
                fila_str = " ".join(textos)
                if any(k in fila_str for k in ["SALDO FINAL", "TOTAL MOV", "* * *"]):
                    break

                mov = self._fila_tabla_a_mov(linea)
                if mov:
                    movimientos.append(mov)

            # Caso 2: linea es dict con texto crudo
            elif isinstance(linea, dict):
                raw = linea.get("_raw", "")
                raw_up = raw.upper()

                if not en_datos:
                    if re.search(r'FECHA.+CARGO.+ABONO', raw_up) or \
                       re.search(r'FECHA.+DESCRIPCI', raw_up):
                        en_datos = True
                    continue

                if any(k in raw_up for k in ["SALDO FINAL", "TOTAL MOV", "* * *"]):
                    break

                mov = self._linea_texto_a_mov(raw)
                if mov:
                    movimientos.append(mov)

        return movimientos

    def _fila_tabla_a_mov(self, fila: list) -> dict | None:
        """Convierte una fila de tabla pdfplumber a movimiento."""
        # Necesitamos al menos 4 columnas
        cols = [str(c or "").strip() for c in fila]
        while len(cols) < 6:
            cols.append("")

        fecha = self.parse_fecha(cols[0])
        if not fecha:
            return None

        descripcion = cols[2] if len(cols) > 2 else cols[1]
        cargo       = self.parse_importe(cols[3] if len(cols) > 3 else "")
        abono       = self.parse_importe(cols[4] if len(cols) > 4 else "")
        saldo       = self.parse_importe(cols[5] if len(cols) > 5 else "")
        nro_op      = cols[1] if len(cols) > 1 else None

        if cargo <= 0 and abono <= 0:
            return None
        if not descripcion or descripcion.upper() in ("CARGO", "ABONO", "FECHA"):
            return None

        tipo    = "cargo" if cargo > 0 else "abono"
        importe = cargo if cargo > 0 else abono

        return {
            "fecha_operacion": fecha,
            "fecha_valor":     fecha,
            "referencia":      nro_op or None,
            "descripcion":     descripcion,
            "tipo":            tipo,
            "importe":         importe,
            "saldo_banco":     saldo or None,
            "moneda":          "PEN",
            "tipo_cambio":     1.0,
        }

    def _linea_texto_a_mov(self, linea: str) -> dict | None:
        """Parsea una línea de texto plano BCP."""
        # Debe comenzar con fecha DD/MM/YYYY
        m = re.match(r'^(\d{2}/\d{2}/\d{4})\s+', linea)
        if not m:
            return None

        fecha = self.parse_fecha(m.group(1))
        if not fecha:
            return None

        # Extraer todos los importes de la línea
        importes = re.findall(r'([\d]{1,3}(?:[,\.]\d{3})*[,\.]\d{2})', linea)
        if len(importes) < 2:
            return None

        saldo   = self.parse_importe(importes[-1])
        importe1 = self.parse_importe(importes[-2]) if len(importes) >= 2 else 0
        importe2 = self.parse_importe(importes[-3]) if len(importes) >= 3 else 0

        # Extraer descripción: todo entre la fecha+nro_op y los importes
        sin_fecha = linea[m.end():]
        # Quitar nro de operación (7-15 dígitos)
        sin_nro = re.sub(r'^\d{7,15}\s+', '', sin_fecha)
        # Quitar importes del final
        desc_raw = re.sub(r'[\d,\.]+\s*$', '', sin_nro)
        desc_raw = re.sub(r'[\d,\.]+\s+', '', desc_raw, count=2)
        descripcion = " ".join(desc_raw.split())

        if not descripcion:
            return None

        # Determinar cargo/abono por posición en la línea
        # En BCP: columna CARGO viene antes que ABONO
        mitad   = len(linea) // 2
        pos_imp = linea.rfind(importes[-2]) if len(importes) >= 2 else 0
        tipo    = "cargo" if pos_imp < mitad else "abono"
        importe = importe1 if importe1 > 0 else importe2

        return {
            "fecha_operacion": fecha,
            "fecha_valor":     fecha,
            "referencia":      None,
            "descripcion":     descripcion[:300],
            "tipo":            tipo,
            "importe":         importe,
            "saldo_banco":     saldo,
            "moneda":          "PEN",
            "tipo_cambio":     1.0,
        }


class BcpExcelParser(BaseParser):
    """
    Parser BCP Excel.
    BCP exporta con cabecera en filas 1-7, datos desde fila 8.
    Columnas: A=Fecha | B=N°Op | C=Descripción | D=Cargo | E=Abono | F=Saldo
    """

    def parsear(self, ruta: str) -> dict:
        res     = self.respuesta_vacia()
        errores = []

        try:
            wb = load_workbook(ruta, read_only=True, data_only=True)
            ws = wb.active

            # Leer cabecera (filas 1-7)
            texto_cab = ""
            for row in ws.iter_rows(min_row=1, max_row=7, values_only=True):
                texto_cab += " ".join(str(c or "") for c in row) + " "

            res["cabecera"] = self._extraer_cabecera_excel(texto_cab)

            # Detectar fila de inicio de datos
            fila_inicio = self._detectar_fila_inicio(ws)

            # Leer movimientos
            movimientos = []
            for row in ws.iter_rows(min_row=fila_inicio, values_only=True):
                fecha = self.parse_fecha(row[0] if row else None)
                if not fecha:
                    continue

                nro_op      = str(row[1] or "").strip() if len(row) > 1 else ""
                descripcion = str(row[2] or "").strip() if len(row) > 2 else ""
                cargo       = self.parse_importe(row[3] if len(row) > 3 else None)
                abono       = self.parse_importe(row[4] if len(row) > 4 else None)
                saldo       = self.parse_importe(row[5] if len(row) > 5 else None)

                if cargo <= 0 and abono <= 0:
                    continue
                if not descripcion:
                    continue

                tipo    = "cargo" if cargo > 0 else "abono"
                importe = cargo if cargo > 0 else abono

                movimientos.append({
                    "fecha_operacion": fecha,
                    "fecha_valor":     fecha,
                    "referencia":      nro_op or None,
                    "descripcion":     descripcion[:300],
                    "tipo":            tipo,
                    "importe":         importe,
                    "saldo_banco":     saldo or None,
                    "moneda":          res["cabecera"]["moneda"],
                    "tipo_cambio":     1.0,
                })

            wb.close()

            if not res["cabecera"].get("periodo") and movimientos:
                res["cabecera"]["periodo"] = self.periodo_desde_fechas(movimientos)

            res["movimientos"]  = movimientos
            res["total_leidos"] = len(movimientos)

        except Exception as e:
            errores.append(f"Error al leer Excel BCP: {str(e)}")

        res["errores"] = errores
        return res

    def _extraer_cabecera_excel(self, texto: str) -> dict:
        t = texto.upper()
        cab = {
            "numero_cuenta": None,
            "periodo":       None,
            "saldo_inicial": None,
            "saldo_final":   None,
            "moneda":        "PEN",
        }
        m = re.search(r'(?:CUENTA|CTA)[^\d]+([\d\s\-]{6,25})', t)
        if m:
            cab["numero_cuenta"] = re.sub(r'[\s\-]', '', m.group(1))[:20]
        m = re.search(r'(\d{2})[/\-](\d{4})', t)
        if m:
            cab["periodo"] = m.group(2) + m.group(1)
        if "DOLAR" in t or "USD" in t:
            cab["moneda"] = "USD"
        return cab

    def _detectar_fila_inicio(self, ws) -> int:
        """Busca la primera fila que tenga una fecha válida en columna A."""
        for i, row in enumerate(ws.iter_rows(min_row=5, max_row=20, values_only=True), start=5):
            if row and self.parse_fecha(row[0]):
                return i
        return 9  # fallback BCP estándar
